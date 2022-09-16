import requests
import re
import os
import time
import datetime
import openpyxl
from bs4 import BeautifulSoup

#省份字典索引
pIndex = dict()
province = ['日期', '本土', '福建', '北京', '河北', '山西', '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽', '江西', '山东', '河南', '湖北', '湖南', '广东', '海南', '四川',
            '贵州', '云南', '陕西', '甘肃', '内蒙古', '青海', '广西', '西藏', '宁夏', '新疆', '天津', '上海', '重庆', '兵团']
for i in range(len(province)):
    pIndex[province[i]] = i

#建表
wb = openpyxl.Workbook()
table1 = wb.create_sheet('本土每日新增确诊')
table2 = wb.create_sheet('本土每日新增无症状')
table3 = wb.create_sheet('港澳台每日累计确诊')
table4 = wb.create_sheet('港澳台每日新增确诊')


#防止requests异常
requests.DEFAULT_RETRIES = 5
s = requests.session()
s.keep_alive = False

#url part
URL1 = "http://www.nhc.gov.cn/xcs/yqtb/list_gzbd.shtml"
URL2 = "http://www.nhc.gov.cn"
URL3 = "/xcs/yqtb/list_gzbd_"
URL4 = ".shtml"
HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "Host": "www.nhc.gov.cn",
    "Referer": "http://www.nhc.gov.cn/",
    "Upgrade-Insecure-Requests": "1",
    "Cookie": "yfx_c_g_u_id_10006654=_ck22090522125813336561160615645; _gscu_2059686908=62442555dbmnfg19; sVoELocvxVW0S=5vYwIaEz4TE1RnPXj_0pHnQNzv_41vuIddaEwJvbuPMwOps.zADXEu5t725V4iLN_S6bfddHgXYPMe25g5VIAgG; sVoELocvxVW0T=53Sv0_bWRei7qqqDko3lLyGgeJ6AGq2G5cMp54a2s_WoH1CjZUscjTWhoEM7FQ88qSG33tkVjjRz21E9F.MgJZl4ouBz6RT08XnOZe5nZtklBxudJwvlYORDvbbMIZrKZatHXiHm5LfuccJ8pKeiomsNBUWumRC7Pll3tdxX02RUeigZmGcd_Yy5Z1SS.YLcPOtMV8WLMLqrBA0uHTXuPVR0KSj7VCp.iZjTphmbxOJ8NQWAe..y85mLJcIAucVC3El4X31Nidmciil0zpbNiRgos7C04jS4snOlBXlU12mL.PRBfK2Rdx9ebd.5wBYVw8o9OZ4PtZGMDRlIdns4TgIGyJtqVYUSj87Ut9Z_FC29q; insert_cookie=91349450; yfx_f_l_v_t_10006654=f_t_1662387178325__r_t_1662643331278__v_t_1662652423426__r_c_2; security_session_verify=f248c756300f2ac90883d087a10f4967",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36"
}
#爬到2020-05-15
Finaldate = datetime.datetime.strptime("2020-05-16", "%Y-%m-%d")

provincePattern = '(河北|山西|辽宁|吉林|黑龙江|江苏|浙江|安徽|福建|江西|山东|河南|湖北|湖南|广东|海南|四川|贵州|云南|陕西|甘肃|内蒙古|青海|台湾|广西|西藏|宁夏|新疆|北京|天津|上海|重庆|香港|澳门|兵团)((\d+)例)*'
#解析每日通报数据

#港澳台连续两天的数据
gatPre = []
gatNow = []

def parseData(date, data):
    print(date)
    newDiagnosed = [0 for x in range(len(province))]
    newAsymptomatic = [0 for x in range(len(province))]
    newDiagnosed[0] = date
    newAsymptomatic[0] = date
    #如果报文有效，obj1.group， obj2.group 格式都为 [当日新增，无关字符， 本土新增， 各省详情]， obj1是确诊，obj2是无症状
    obj1 = re.search('报告新增确诊病例.*?(\d+)例(.*?)本土病例(\d+)例(（(.*?)）)', data)
    obj2 = re.search('新增无症状感染者(\d+)例(.*?)本土.*?(\d+)例(（(.*?)）)', data)
    obj3 = re.search('香港.+?(\d+)例', data)
    obj4 = re.search('澳门.+?(\d+)例', data)
    obj5 = re.search('台湾.+?(\d+)例', data)
    #无效通报页
    if obj1 == None and obj2 == None and obj3 == None and obj4 == None and obj5 == None:
        return
    #2022-02-21特殊报文格式，特殊处理
    if obj1 != None and date != '2022-02-21':
        print('新增本土确诊病例', obj1.group(3), '例', obj1.group(4))
        newDiagnosed[1] = int(obj1.group(3))
        provinceDate = re.findall(provincePattern, obj1.group(4))
        if(len(provinceDate) == 1):
            print(provinceDate[0][0], obj1.group(3))
            newDiagnosed[pIndex[provinceDate[0][0]]] = int(obj1.group(3))
        else:
            #tol = 0
            for x in provinceDate:
                if newDiagnosed[pIndex[x[0]]] != 0 or x[1] == x[2]:
                    continue
                #tol += int(x[2])
                print(x[0], x[2], end= ' ')
                newDiagnosed[pIndex[x[0]]] = int(x[2])
            #if tol != int(obj1.group(3)):
            #   print('确诊不匹配')
            #   while 1:
            #       pass
    elif obj1 == None:
        print('无新增本土确诊病例')
    else:
        newDiagnosed = ['2022-02-21', 59, 0, 4, 0, 0, 6, 0, 0, 12, 0, 0, 0, 0, 0, 4, 0, 3, 0, 5, 0, 5, 0, 0, 20, 0, 0,0, 0, 0, 0, 0, 0, 0]
    # 2021-10-06为特殊文件格式， 特殊处理
    if obj2 != None and date != '2021-10-06':
        print('\n新增本土无症状感染者', obj2.group(3), '例', obj2.group(4))
        newAsymptomatic[1] = int(obj2.group(3))
        provinceDate = re.findall(provincePattern, obj2.group(4))
        if len(provinceDate) == 1:
            print(provinceDate[0][0], obj2.group(3))
            newAsymptomatic[pIndex[provinceDate[0][0]]] = int(obj2.group(3))
        else:
            #tol = 0
            for x in provinceDate:
                if newAsymptomatic[pIndex[x[0]]] != 0 or x[1] == x[2]:
                    continue
                #tol += int(x[2])
                print(x[0], x[2], end=' ')
                newAsymptomatic[pIndex[x[0]]] = int(x[2])
            #if tol != int(obj2.group(3)):
            #    print('无症状不匹配')
            #    while 1:
            #        pass
    elif obj2 == None:
        print('无新增本土无症状感染者')
    else:
        newAsymptomatic = ['2021-10-06', 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,0, 0, 0, 0, 0, 0, 2]
    print('\n香港累计确诊', obj3.group(1), '例')
    print('澳门累计确诊', obj4.group(1), '例')
    print('台湾累计确诊', obj5.group(1), '例')
    #港澳台今日累计新增
    gat = [date, int(obj3.group(1)), int(obj4.group(1)), int(obj5.group(1))]
    table1.append(newDiagnosed)
    table2.append(newAsymptomatic)
    table3.append(gat)
    global gatPre
    global gatNow
    print(newDiagnosed)
    print(newAsymptomatic)
    print(gat)
    if len(gatPre) == 0:
        gatPre = gat
    else:
        gatNow = gat
        for i in (1, 2, 3):
            gatPre[i] -= gatNow[i]
        table4.append(gatPre)
        print("港澳台每日新增：", end='')
        print(gatPre)
        gatPre = gatNow

#获得每日通报报文
def getData(date, link):
    while True:
        time.sleep(0.5)
        #爬每日通报
        html = requests.get(url=link, headers=HEADERS).text
        #解析每日通报
        sp = BeautifulSoup(html, 'lxml')
        #反爬，重来
        if re.search('w3.org', html):
            continue
        #反扒，重来
        try:
            lst = sp.find('div', attrs={"id": "xw_box"}).find_all("p")
        except AttributeError:
            continue
        #保存报文
        s = ""
        for word in lst:
            s += word.text
        #解析报文, 如果parseData抛出typeerror， 说明爬虫爬到错误页
        while True:
            try:
                parseData(date, s)
            except TypeError:
                continue
            break
        print(date, 'done')
        break

#获得每日通报url
def getUrl():
    i = 1
    while True:
        if i == 1:
            url = URL1
        else:
            url = URL2 + URL3 + str(i) + URL4
        time.sleep(0.5)
        #爬目录页
        response = requests.get(url=url, headers=HEADERS).text
        if re.search('w3.org', response):
            #反爬， 重来
            pass
        else:
            #成功，解析目录页
            soup = BeautifulSoup(response, 'lxml')
            #解析目录页目录项，在html li标签中
            for node in soup.find_all('li'):
                link = URL2 + node.a['href']
                date = node.span.text
                dateObj = datetime.datetime.strptime(date, "%Y-%m-%d")
                #统计到2020-05-15
                if (Finaldate - dateObj).days > 0:
                    return
                td = datetime.timedelta(days=-1)
                dateObj = dateObj + td
                date = dateObj.strftime("%Y-%m-%d")
                #获得当天通报数据
                getData(date, link)
            i += 1

def yqtbCrawler():
    if os.path.exists('疫情通报.xlsx') == False:
        table1.append(province)
        table2.append(province)
        table3.append(['日期', '香港', '澳门', '台湾'])
        table4.append(['日期', '香港', '澳门', '台湾'])
        wb.remove(wb["Sheet"])
        getUrl()
        wb.save('疫情通报.xlsx')
        wb.close()
if __name__ == '__main__':
    yqtbCrawler()
